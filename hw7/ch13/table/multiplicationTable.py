#!/usr/bin/env python3
''' Mulitplication table maker

Author: Cate Lukner 
'''
import argparse
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font

def create_multiplication_table(N):

    '''Given a number N, create an NxN multiplication table in an
    Excel spreadsheet. This sheet's filename should be "output.xlsx"

    Args:
      N (int): size of multiplication table

    Returns:
      None

    '''

    # Open a new workbook
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    
    # Ensure N is an integer
    N = int(N)

    # Label the NxN table
    for i in range (2, 2+N):
        sheet.cell(row=i, column=1).value = i-1
        sheet.cell(row=i, column=1).font = Font(bold=True) 
        sheet.cell(row=1, column=i).value = i-1
        sheet.cell(row=1, column=i).font = Font(bold=True)
    
    # Multiply NxN in each cell
    for i in range(2, sheet.max_row + 1):
        for j in range(2, sheet.max_column + 1):
            x = sheet.cell(row=i, column=1).value
            y = sheet.cell(row=1, column=j).value
            sheet.cell(row=i, column=j).value = x * y

    # Save as output.xlsx
    print(f'Created {N}X{N} table as output.xlsx')
    wb.save('output.xlsx')

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('N', help='size of multiplication table')

    args = parser.parse_args()
    
    N = int(args.N)
    create_multiplication_table(args.N)

if __name__=='__main__':
    main()
